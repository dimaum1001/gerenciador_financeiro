import uuid
from typing import Optional, List, Tuple, Dict, Any, Set
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID
from zipfile import BadZipFile

import unicodedata
from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    status,
    Query,
    UploadFile,
    File,
    Form,
)
from openpyxl import load_workbook
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, desc, func, extract

from app.core.config import settings
from app.core.deps import get_current_user, get_current_non_demo_user, get_db
from app.models.user import User

from app.models.transaction import (
    Transaction,
    TransactionType,
    TransactionStatus,
    PaymentMethod,
)
from app.models.account import Account, AccountType
from app.models.category import Category, CategoryType
from app.schemas.transaction import (
    TransactionCreate,
    TransactionUpdate,
    TransactionResponse,
    TransactionListResponse,
    TransactionSummary,
    TransactionImportResult,
)
from app.utils.locale_mapper import (
    transaction_type_mapper,
    transaction_status_mapper,
    payment_method_mapper,
    account_type_mapper,
)
from app.utils.pagination import paginate_query

router = APIRouter()


def _transaction_query(db: Session, current_user: User):
    return db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.is_demo_data.is_(current_user.is_demo),
    )


def _account_query(db: Session, current_user: User):
    return db.query(Account).filter(
        Account.user_id == current_user.id,
        Account.is_demo_data.is_(current_user.is_demo),
    )


def _category_query(db: Session, current_user: User):
    return db.query(Category).filter(
        Category.user_id == current_user.id,
        Category.is_demo_data.is_(current_user.is_demo),
    )


def _row_is_empty(row: Tuple[Any, ...]) -> bool:
    """
    Verifica se a linha do Excel est�� vazia (todos os valores nulos ou strings vazias)
    """
    return all(
        cell is None or (isinstance(cell, str) and not cell.strip())
        for cell in row
    )




def _normalize_lookup_key(value: str) -> str:
    '''Remove acentos e normaliza o texto para facilitar comparacoes.'''
    if not value:
        return ''
    normalized = unicodedata.normalize('NFKD', str(value))
    base_text = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return base_text.strip().lower()


def _ensure_account_for_import(
    account_name: str,
    normalized_name: str,
    account_type_value: Optional[str],
    default_account: Optional[Account],
    account_lookup_by_name: Dict[str, Account],
    account_lookup_by_id: Dict[str, Account],
    current_user: User,
    db: Session,
    dry_run: bool,
) -> Account:
    """
    Cria uma nova conta para o usuario quando o nome informado ainda nao existe.
    """
    display_name = (account_name or "").strip()
    if not display_name or not normalized_name:
        raise ValueError("Informe um valor valido na coluna 'conta_nome'.")

    if normalized_name in account_lookup_by_name:
        return account_lookup_by_name[normalized_name]

    if not default_account:
        raise ValueError(
            "Selecione uma conta padrao no sistema para permitir a criacao de contas novas."
        )

    account_type = default_account.tipo or AccountType.CHECKING
    if account_type_value:
        try:
            mapped_type = account_type_mapper.to_enum(account_type_value)
        except ValueError:
            raise ValueError(
                "Tipo de conta invalido. Use dinheiro, conta_corrente, poupanca, "
                "cartao_credito, investimento ou outros."
            )
        if mapped_type:
            account_type = mapped_type

    moeda = default_account.moeda or "BRL"

    new_account = Account(
        id=uuid.uuid4(),
        nome=display_name[:100],
        tipo=account_type,
        user_id=current_user.id,
        is_demo_data=current_user.is_demo,
        moeda=moeda,
        ativo=True,
    )

    if not dry_run:
        db.add(new_account)

    account_lookup_by_name[normalized_name] = new_account
    account_lookup_by_id[str(new_account.id).strip().lower()] = new_account
    return new_account


def _resolve_account_from_row(
    row_data: Dict[str, Any],
    default_account: Optional[Account],
    account_lookup_by_id: Dict[str, Account],
    account_lookup_by_name: Dict[str, Account],
    duplicated_account_names: Set[str],
    current_user: User,
    db: Session,
    dry_run: bool,
) -> Account:
    '''Determina a conta da transacao usando o ID ou o nome informado na planilha.'''
    conta_id_raw = (
        row_data.get('conta_id')
        or row_data.get('account_id')
        or row_data.get('conta_destino_id')
    )
    if conta_id_raw:
        conta_id_key = str(conta_id_raw).strip().lower()
        account = account_lookup_by_id.get(conta_id_key)
        if not account:
            raise ValueError(
                f"Conta com ID '{conta_id_raw}' nao encontrada para este usuario"
            )
        return account

    conta_nome_raw = (
        row_data.get('conta_nome')
        or row_data.get('account_name')
        or row_data.get('conta')
    )
    if conta_nome_raw:
        normalized_name = _normalize_lookup_key(str(conta_nome_raw))
        if normalized_name in duplicated_account_names:
            raise ValueError(
                f"Mais de uma conta utiliza o nome '{conta_nome_raw}'. "
                "Informe o ID usando a coluna 'conta_id'."
            )
        account = account_lookup_by_name.get(normalized_name)
        if not account:
            account_type_value = (
                row_data.get("conta_tipo")
                or row_data.get("account_type")
                or row_data.get("tipo_conta")
            )
            account = _ensure_account_for_import(
                account_name=str(conta_nome_raw),
                normalized_name=normalized_name,
                account_type_value=account_type_value,
                default_account=default_account,
                account_lookup_by_name=account_lookup_by_name,
                account_lookup_by_id=account_lookup_by_id,
                current_user=current_user,
                db=db,
                dry_run=dry_run,
            )
        return account

    if default_account:
        return default_account

    raise ValueError(
        "Informe a conta de destino usando as colunas 'conta_nome' ou 'conta_id'."
    )


def _parse_date_value(value: Any, field_name: str, required: bool = False) -> Optional[date]:
    """
    Converte valores aceitos em datas (date, datetime ou string ISO).
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise ValueError(f"Campo '{field_name}' �� obrigat��rio")
        return None
    
    if isinstance(value, datetime):
        return value.date()
    
    if isinstance(value, date):
        return value
    
    if isinstance(value, str):
        text = value.strip()
        try:
            return date.fromisoformat(text)
        except ValueError:
            raise ValueError(f"Campo '{field_name}' deve estar no formato AAAA-MM-DD")
    
    raise ValueError(f"Campo '{field_name}' possui formato de data inv��lido")


def _parse_decimal_value(value: Any) -> Decimal:
    """
    Normaliza valores num��ricos vindo da planilha.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError("Campo 'valor' �� obrigat��rio")
    
    decimal_value: Decimal
    
    if isinstance(value, Decimal):
        decimal_value = value
    elif isinstance(value, (int, float)):
        decimal_value = Decimal(str(value))
    elif isinstance(value, str):
        text = (
            value.replace("R$", "")
            .replace(" ", "")
            .replace("\u00a0", "")
            .strip()
        )
        if not text:
            raise ValueError("Campo 'valor' �� obrigat��rio")
        
        comma_count = text.count(",")
        dot_count = text.count(".")
        
        if comma_count and dot_count:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif comma_count and not dot_count:
            if comma_count == 1:
                text = text.replace(",", ".")
            else:
                text = text.replace(",", "")
        elif dot_count > 1:
            text = text.replace(".", "")
        
        try:
            decimal_value = Decimal(text)
        except InvalidOperation:
            raise ValueError("Campo 'valor' possui formato num��rico inv��lido")
    else:
        raise ValueError("Campo 'valor' possui formato n��o suportado")
    
    if decimal_value <= 0:
        raise ValueError("Campo 'valor' deve ser maior que zero")
    
    try:
        return decimal_value.quantize(Decimal("0.01"))
    except InvalidOperation:
        raise ValueError("Campo 'valor' deve ter no m��ximo duas casas decimais")


def _extract_tags(raw_value: Any) -> List[str]:
    if raw_value in (None, "", []):
        return []
    
    if isinstance(raw_value, str):
        parts = raw_value.split(",")
    elif isinstance(raw_value, (list, tuple, set)):
        parts = [str(item) for item in raw_value]
    else:
        return []
    
    return [tag.strip() for tag in parts if tag and tag.strip()]


def _ensure_category_for_import(
    category_name: str,
    transaction_type: TransactionType,
    category_lookup: Dict[str, Category],
    current_user: User,
    db: Session,
    dry_run: bool,
) -> Category:
    """Retorna categoria existente ou cria uma nova automaticamente durante importação."""
    normalized_name = (category_name or "").strip()
    if not normalized_name:
        return None

    lookup_key = normalized_name.lower()
    existing = category_lookup.get(lookup_key)
    if existing:
        return existing

    category_type = (
        CategoryType.INCOME if transaction_type == TransactionType.INCOME else CategoryType.EXPENSE
    )

    new_category = Category(
        id=uuid.uuid4(),
        nome=normalized_name,
        tipo=category_type,
        user_id=current_user.id,
        is_demo_data=current_user.is_demo,
        ativo=True,
    )

    if not dry_run:
        db.add(new_category)

    category_lookup[lookup_key] = new_category
    return new_category



def _prepare_transaction_from_row(
    row_data: Dict[str, Any],
    line_number: int,
    category_lookup: Dict[str, Category],
    current_user: User,
    db: Session,
    dry_run: bool,
    default_account: Account,
    account_lookup_by_id: Dict[str, Account],
    account_lookup_by_name: Dict[str, Account],
    duplicated_account_names: Set[str],
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Converte uma linha da planilha em dados prontos para criar uma transacao."""
    tipo_raw = str(row_data.get("tipo") or "").strip()
    if not tipo_raw:
        raise ValueError("Campo 'tipo' e obrigatorio")

    try:
        tipo = transaction_type_mapper.to_enum(tipo_raw)
    except ValueError:
        raise ValueError("Tipo invalido. Use receita, despesa ou transferencia (ou income/expense/transfer).")

    if tipo == TransactionType.TRANSFER:
        raise ValueError("Importacao de transferencias nao e suportada neste modelo")

    data_lancamento = _parse_date_value(row_data.get("data_lancamento"), "data_lancamento", required=True)
    descricao = str(row_data.get("descricao") or "").strip()
    if not descricao:
        raise ValueError("Campo 'descricao' e obrigatorio")
    descricao = descricao[:255]

    valor = _parse_decimal_value(row_data.get("valor"))

    moeda = str(row_data.get("moeda") or "BRL").strip().upper() or "BRL"
    moeda = moeda[:3]

    status_raw = str(row_data.get("status") or TransactionStatus.PENDING.value).strip()
    try:
        status = transaction_status_mapper.to_enum(status_raw)
    except ValueError:
        raise ValueError("Status invalido. Use pendente, compensada ou conciliada (ou pending/cleared/reconciled)")

    payment_method_raw = str(row_data.get("payment_method") or "").strip()
    payment_method = None
    if payment_method_raw:
        try:
            payment_method = payment_method_mapper.to_enum(payment_method_raw)
        except ValueError:
            raise ValueError("Metodo de pagamento invalido")

    categoria_nome = row_data.get("categoria_nome")
    category = None
    if categoria_nome:
        categoria_key = str(categoria_nome).strip().lower()
        category = category_lookup.get(categoria_key)
        if not category:
            category = _ensure_category_for_import(
                category_name=str(categoria_nome),
                transaction_type=tipo,
                category_lookup=category_lookup,
                current_user=current_user,
                db=db,
                dry_run=dry_run,
            )

    account = _resolve_account_from_row(
        row_data=row_data,
        default_account=default_account,
        account_lookup_by_id=account_lookup_by_id,
        account_lookup_by_name=account_lookup_by_name,
        duplicated_account_names=duplicated_account_names,
        current_user=current_user,
        db=db,
        dry_run=dry_run,
    )

    data_competencia = _parse_date_value(row_data.get("data_competencia"), "data_competencia")
    tags = _extract_tags(row_data.get("tags"))
    observacoes_raw = row_data.get("observacoes")
    observacoes = str(observacoes_raw).strip() if observacoes_raw not in (None, "") else None

    payload = {
        "tipo": tipo,
        "valor": valor,
        "moeda": moeda,
        "data_lancamento": data_lancamento,
        "descricao": descricao,
        "status": status,
        "payment_method": payment_method,
        "category_id": category.id if category else None,
        "data_competencia": data_competencia,
        "tags": tags,
        "observacoes": observacoes,
        "account_id": account.id,
    }

    preview_entry = {
        "linha": line_number,
        "tipo": tipo.value,
        "descricao": descricao,
        "valor": float(valor),
        "moeda": moeda,
        "data_lancamento": data_lancamento.isoformat(),
        "categoria": category.nome if category else None,
        "status": status.value,
        "conta": account.nome,
    }

    if payment_method:
        preview_entry["payment_method"] = payment_method.value

    return payload, preview_entry

@router.get("", include_in_schema=False, response_model=TransactionListResponse)
@router.get("/", response_model=TransactionListResponse)
async def list_transactions(
    skip: int = 0,
    limit: int = 50,
    tipo: Optional[str] = None,
    categoria_id: Optional[str] = None,
    conta_origem_id: Optional[str] = None,
    conta_destino_id: Optional[str] = None,
    status: Optional[str] = None,
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    valor_min: Optional[float] = None,
    valor_max: Optional[float] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Listar transações do usuário com filtros avançados"""
    query = _transaction_query(db, current_user).options(
        joinedload(Transaction.category),
        joinedload(Transaction.account),
        joinedload(Transaction.transfer_account)
    )
    
    # Aplicar filtros
    if tipo:
        try:
            tipo_enum = transaction_type_mapper.to_enum(tipo)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Tipo de transação inválido",
            )
        query = query.filter(Transaction.tipo == tipo_enum)
    
    if categoria_id:
        query = query.filter(Transaction.category_id == categoria_id)
    
    if conta_origem_id:
        query = query.filter(Transaction.account_id == conta_origem_id)
    
    if conta_destino_id:
        query = query.filter(Transaction.transfer_account_id == conta_destino_id)
    
    if status:
        try:
            status_enum = transaction_status_mapper.to_enum(status)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Status de transação inválido",
            )
        query = query.filter(Transaction.status == status_enum)
    
    if data_inicio:
        query = query.filter(Transaction.data_lancamento >= data_inicio)
    
    if data_fim:
        query = query.filter(Transaction.data_lancamento <= data_fim)
    
    if valor_min is not None:
        query = query.filter(Transaction.valor >= Decimal(str(valor_min)))
    
    if valor_max is not None:
        query = query.filter(Transaction.valor <= Decimal(str(valor_max)))
    
    if search:
        query = query.filter(
            or_(
                Transaction.descricao.ilike(f"%{search}%"),
                Transaction.observacoes.ilike(f"%{search}%")
            )
        )
    
    transactions, total = paginate_query(
        query.order_by(desc(Transaction.data_lancamento), desc(Transaction.criado_em)),
        skip=skip,
        limit=limit,
    )
    
    return TransactionListResponse(
        transactions=transactions,
        total=total,
        skip=skip,
        limit=limit
    )

@router.post("", include_in_schema=False, response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
async def create_transaction(
    transaction_data: TransactionCreate,
    current_user: User = Depends(get_current_non_demo_user),
    db: Session = Depends(get_db)
):
    """Criar nova transação"""
    # Validar contas
    if transaction_data.account_id:
        conta_origem = (
            _account_query(db, current_user)
            .filter(Account.id == transaction_data.account_id)
            .first()
        )
        if not conta_origem:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Conta de origem não encontrada"
            )
    
    if transaction_data.transfer_account_id:
        conta_destino = (
            _account_query(db, current_user)
            .filter(Account.id == transaction_data.transfer_account_id)
            .first()
        )
        if not conta_destino:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Conta de destino não encontrada"
            )
    
    # Validar categoria
    if transaction_data.category_id:
        categoria = (
            _category_query(db, current_user)
            .filter(Category.id == transaction_data.category_id)
            .first()
        )
        if not categoria:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Categoria não encontrada"
            )
    
    # Criar transação
    transaction = Transaction(
        **transaction_data.model_dump(),
        user_id=current_user.id,
        is_demo_data=current_user.is_demo
    )
    
    db.add(transaction)
    db.commit()
    db.refresh(transaction)
    
    # Atualizar saldos das contas
    await _update_account_balances(transaction, db)
    
    return transaction


@router.post("/import", response_model=TransactionImportResult)
async def import_transactions_from_template(
    account_id: UUID = Form(...),
    dry_run: bool = Form(True),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_non_demo_user),
    db: Session = Depends(get_db),
):
    """Importar transa����es a partir do template Excel."""
    if not settings.enable_import:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Importa��o est�� desabilitada neste ambiente",
        )
    
    user_accounts = _account_query(db, current_user).all()
    conta_destino = next((account for account in user_accounts if account.id == account_id), None)
    if not conta_destino:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Conta de destino nao encontrada",
        )
    
    account_lookup_by_id: Dict[str, Account] = {
        str(account.id).strip().lower(): account
        for account in user_accounts
    }
    account_lookup_by_name: Dict[str, Account] = {}
    duplicated_account_names: Set[str] = set()
    for account in user_accounts:
        if not account.nome:
            continue
        normalized_name = _normalize_lookup_key(account.nome)
        if not normalized_name:
            continue
        if normalized_name in account_lookup_by_name:
            duplicated_account_names.add(normalized_name)
            continue
        account_lookup_by_name[normalized_name] = account
    
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Envie um arquivo .xlsx gerado a partir do modelo disponibilizado",
        )
    
    contents = await file.read()
    if not contents:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Arquivo recebido est�� vazio",
        )
    
    try:
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
    except BadZipFile:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Arquivo corrompido ou formato inv��lido. Gere o modelo novamente.",
        )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="N��o foi poss��vel ler o arquivo enviado",
        )
    
    if "Transacoes" not in workbook.sheetnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A planilha deve conter a aba 'Transacoes'",
        )
    
    sheet = workbook["Transacoes"]
    try:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Planilha sem cabe��alho. Use o modelo atualizado.",
        )
    
    header_map: List[Optional[str]] = []
    for header_cell in header_row:
        if isinstance(header_cell, str):
            normalized = header_cell.strip().lower()
            header_map.append(normalized if normalized else None)
        elif header_cell is None:
            header_map.append(None)
        else:
            header_map.append(str(header_cell).strip().lower())
    
    if not any(header_map):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cabe��alho n��o encontrado. Verifique o modelo utilizado.",
        )
    
    required_columns = {"tipo", "data_lancamento", "descricao", "valor"}
    header_set = {name for name in header_map if name}
    missing_columns = required_columns - header_set
    if missing_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Colunas obrigat��rias ausentes: {', '.join(sorted(missing_columns))}",
        )
    
    category_lookup = {
        (category.nome or "").strip().lower(): category
        for category in _category_query(db, current_user).all()
        if category.nome
    }
    
    total_rows = 0
    processed_rows = 0
    errors: List[str] = []
    preview: List[Dict[str, Any]] = []
    rows_to_create: List[Dict[str, Any]] = []
    
    for line_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row is None or _row_is_empty(row):
            continue
        
        total_rows += 1
        row_dict: Dict[str, Any] = {}
        for idx, value in enumerate(row):
            if idx >= len(header_map):
                break
            header_name = header_map[idx]
            if header_name:
                row_dict[header_name] = value
        
        try:
            payload, preview_entry = _prepare_transaction_from_row(
                row_dict,
                line_number,
                category_lookup,
                current_user,
                db,
                dry_run,
                conta_destino,
                account_lookup_by_id,
                account_lookup_by_name,
                duplicated_account_names,
            )
        except ValueError as exc:
            errors.append(f"Linha {line_number}: {exc}")
            continue
        
        rows_to_create.append(payload)
        processed_rows += 1
        if len(preview) < 5:
            preview.append(preview_entry)
    
    created_transactions: List[Transaction] = []
    if rows_to_create and not dry_run:
        try:
            for payload in rows_to_create:
                transaction = Transaction(
                    **payload,
                    user_id=current_user.id,
                    is_demo_data=current_user.is_demo,
                )
                db.add(transaction)
                created_transactions.append(transaction)
            
            db.commit()
            
            for transaction in created_transactions:
                db.refresh(transaction)
                await _update_account_balances(transaction, db)
        except Exception:
            db.rollback()
            raise
    
    return TransactionImportResult(
        total_linhas=total_rows,
        linhas_processadas=processed_rows,
        linhas_com_erro=len(errors),
        transacoes_criadas=len(created_transactions) if not dry_run else 0,
        erros=errors,
        preview=preview,
    )
@router.get("/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(
    transaction_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Obter transação específica"""
    transaction = (
        _transaction_query(db, current_user)
        .options(
            joinedload(Transaction.category),
            joinedload(Transaction.account),
            joinedload(Transaction.transfer_account)
        )
        .filter(Transaction.id == transaction_id)
        .first()
    )
    
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transação não encontrada"
        )
    
    return transaction

@router.put("/{transaction_id}", response_model=TransactionResponse)
async def update_transaction(
    transaction_id: str,
    transaction_data: TransactionUpdate,
    current_user: User = Depends(get_current_non_demo_user),
    db: Session = Depends(get_db)
):
    """Atualizar transação"""
    transaction = _transaction_query(db, current_user).filter(Transaction.id == transaction_id).first()
    
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transação não encontrada"
        )
    
    # Reverter saldos antigos
    await _revert_account_balances(transaction, db)
    
    # Atualizar campos
    update_data = transaction_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(transaction, field, value)
    
    db.commit()
    db.refresh(transaction)
    
    # Aplicar novos saldos
    await _update_account_balances(transaction, db)
    
    return transaction

@router.delete("/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_transaction(
    transaction_id: str,
    current_user: User = Depends(get_current_non_demo_user),
    db: Session = Depends(get_db)
):
    """Excluir transação"""
    transaction = _transaction_query(db, current_user).filter(Transaction.id == transaction_id).first()
    
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transação não encontrada"
        )
    
    # Reverter saldos
    await _revert_account_balances(transaction, db)
    
    # Excluir transação
    db.delete(transaction)
    db.commit()

@router.get("/summary/monthly")
async def get_monthly_summary(
    year: int = Query(default=datetime.now().year),
    month: int = Query(default=datetime.now().month),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Obter resumo mensal de transações"""
    demo_condition = Transaction.is_demo_data.is_(current_user.is_demo)
    # Receitas
    receitas = db.query(func.sum(Transaction.valor)).filter(
        and_(
            Transaction.user_id == current_user.id,
            demo_condition,
            Transaction.tipo == TransactionType.INCOME.value,
            extract('year', Transaction.data_lancamento) == year,
            extract('month', Transaction.data_lancamento) == month
        )
    ).scalar() or 0
    
    # Despesas
    despesas = db.query(func.sum(Transaction.valor)).filter(
        and_(
            Transaction.user_id == current_user.id,
            demo_condition,
            Transaction.tipo == TransactionType.EXPENSE.value,
            extract('year', Transaction.data_lancamento) == year,
            extract('month', Transaction.data_lancamento) == month
        )
    ).scalar() or 0
    
    # Transferências
    transferencias = db.query(func.sum(Transaction.valor)).filter(
        and_(
            Transaction.user_id == current_user.id,
            demo_condition,
            Transaction.tipo == TransactionType.TRANSFER.value,
            extract('year', Transaction.data_lancamento) == year,
            extract('month', Transaction.data_lancamento) == month
        )
    ).scalar() or 0
    
    return {
        "year": year,
        "month": month,
        "receitas": float(receitas),
        "despesas": float(abs(despesas)),
        "transferencias": float(transferencias),
        "saldo": float(receitas + despesas),  # despesas já são negativas
        "economia": float(receitas + despesas) if receitas > 0 else 0
    }

@router.get("/summary/by-category")
async def get_summary_by_category(
    tipo: str = Query(..., regex="^(income|expense)$"),
    year: int = Query(default=datetime.now().year),
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Obter resumo por categoria"""
    demo_condition = Transaction.is_demo_data.is_(current_user.is_demo)
    query = db.query(
        Category.nome.label('categoria'),
        Category.cor.label('cor'),
        func.sum(Transaction.valor).label('total'),
        func.count(Transaction.id).label('quantidade')
    ).join(
        Transaction, Transaction.category_id == Category.id
    ).filter(
        and_(
            Transaction.user_id == current_user.id,
            demo_condition,
            Transaction.tipo == tipo,
            extract('year', Transaction.data_lancamento) == year
        )
    ).filter(Category.is_demo_data.is_(current_user.is_demo))
    
    if month:
        query = query.filter(extract('month', Transaction.data_lancamento) == month)
    
    results = query.group_by(Category.id, Category.nome, Category.cor).all()
    
    total_geral = sum(float(r.total) for r in results)
    
    return [
        {
            "categoria": r.categoria,
            "cor": r.cor,
            "valor": float(abs(r.total)),
            "quantidade": r.quantidade,
            "percentual": (float(abs(r.total)) / total_geral * 100) if total_geral > 0 else 0
        }
        for r in results
    ]

def _apply_account_balances(transaction: Transaction, db: Session, multiplier: int) -> None:
    """Aplica variação de saldo nas contas envolvidas.

    Atualmente os saldos são calculados dinamicamente a partir das transações.
    Portanto esta função é um no-op mantido apenas para compatibilidade.
    """
    return None


async def _update_account_balances(transaction: Transaction, db: Session) -> None:
    """Aplica o efeito da transação nas contas."""
    _apply_account_balances(transaction, db, multiplier=1)


async def _revert_account_balances(transaction: Transaction, db: Session) -> None:
    """Reverte o efeito da transação nas contas."""
    _apply_account_balances(transaction, db, multiplier=-1)

